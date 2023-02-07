import csv

import openpyxl
from django.http import HttpResponse
from django.views.decorators.cache import cache_page
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils.decorators import method_decorator


from training.api.serializers import (
    TagSerializer,
    TrainingDataListSerializer,
    TrainingDataSerializer,
    TrainingDataTagSerializer,
    TrainingDataUploadSerializer,
)
from training.models import Tag, TrainingData


class TrainingDataViewSet(viewsets.ModelViewSet):
    queryset = TrainingData.objects.all()

    def get_serializer_class(self):
        if self.request.method == "GET":
            return TrainingDataListSerializer
        else:
            return TrainingDataSerializer

    @action(
        methods=["post"], detail=True, url_path="tag", serializer_class=TagSerializer
    )
    def add_tag(self, request, pk=None):
        instance = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tag = serializer.save()
        instance.tags.add(tag)
        instance.save()
        return Response({"message": "Tag added successfully"})

    @method_decorator(cache_page(60 * 15))
    @action(methods=["GET"], detail=False)
    def aspects(self, request):
        aspects = Tag.objects.values_list("aspect", flat=True).distinct()
        return Response({"aspects": aspects})

    @method_decorator(cache_page(60 * 15))
    @action(methods=["GET"], detail=False)
    def sentiments(self, request):
        sentiments = Tag.objects.values_list("sentiment", flat=True).distinct()
        return Response({"sentiments": sentiments})

    @action(
        methods=["post"], detail=False, serializer_class=TrainingDataUploadSerializer
    )
    def upload(self, request):
        file = request.FILES["file"]

        # Check if the file is a CSV or XLSX file
        if file.name.endswith(".csv"):
            # Read the CSV file and validate the data
            decoded_file = file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)
            data = []

            for row in reader:
                tags = []
                updated_row = {"text": row.get("text"), "tags": eval(row.get("tags"))}
                data.append(updated_row)

        elif file.name.endswith(".xlsx"):
            # Read the XLSX file and validate the data
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook[workbook.sheetnames[0]]
            data = [
                {"text": row[0], "tags": eval(row[1])}
                for index, row in enumerate(worksheet.iter_rows(values_only=True))
                if index != 0
            ]

        else:
            raise Exception("Invalid file format")

        serializer = TrainingDataTagSerializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)

        # Save the validated data to the database
        self.perform_create(serializer)

        return Response(
            {"data": "Data upload successful"}, status=status.HTTP_201_CREATED
        )

    @action(methods=["get"], detail=False)
    def download(self, request):
        file_type = request.query_params.get("format", "csv")
        data = self.get_queryset().values("text", "tags")

        if file_type == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="training_data.csv"'

            writer = csv.writer(response)
            writer.writerow(["text", "tags"])
            for row in data:
                tag = Tag.objects.get(id=row["tags"])
                row["tags"] = TagSerializer(tag).data
                writer.writerow([row["text"], row["tags"]])

            return response
        elif file_type == "xlsx":
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response[
                "Content-Disposition"
            ] = 'attachment; filename="training_data.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(["text", "tags"])
            for row in data:
                tag = Tag.objects.get(id=row["tags"])
                row["tags"] = TagSerializer(tag).data
                worksheet.append([row["text"], row["tags"]])

            workbook.save(response)

            return response
        else:
            raise Exception("Invalid file format")
